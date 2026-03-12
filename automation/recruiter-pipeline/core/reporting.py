from __future__ import annotations

import re
from datetime import datetime
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .models import CandidateResult, MailItem

THIN = Side(style='thin', color='D9E2F2')
HEADER_FILL = PatternFill('solid', fgColor='1F4E78')
SUBHEADER_FILL = PatternFill('solid', fgColor='D9EAF7')
PASS_FILL = PatternFill('solid', fgColor='EAF7EA')
SKIP_FILL = PatternFill('solid', fgColor='F8F9FB')
WARN_FILL = PatternFill('solid', fgColor='FFF4E5')


MOBILE_PATTERNS = [
    re.compile(r'(?<!\d)(1[3-9]\d{9})(?!\d)'),
    re.compile(r'(?<!\d)(?:86[- ]?)?(1[3-9]\d{9})(?!\d)'),
]
EMAIL_PATTERN = re.compile(r'[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}')
YEARS_PATTERN = re.compile(r'(\d+)\s*年')



def _find_mobile(text: str) -> str:
    for pattern in MOBILE_PATTERNS:
        match = pattern.search(text)
        if match:
            return match.group(1)
    return ''



def _find_email(text: str) -> str:
    match = EMAIL_PATTERN.search(text)
    return match.group(0) if match else ''



def _find_years(text: str) -> str:
    years = [int(x) for x in YEARS_PATTERN.findall(text)]
    return str(max(years)) if years else ''



def _attachment_names(documents: Iterable[dict[str, str]]) -> str:
    names = []
    for doc in documents:
        file_name = str(doc.get('file') or '').strip()
        if file_name:
            names.append(file_name)
    return '\n'.join(names)



def _style_title(cell) -> None:
    cell.font = Font(name='Aptos Display', bold=True, color='FFFFFF', size=16)
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal='center', vertical='center')



def _style_header_row(row) -> None:
    for cell in row:
        cell.font = Font(name='Aptos', bold=True, color='FFFFFF')
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)



def _style_data_row(row, fill=None) -> None:
    for cell in row:
        cell.alignment = Alignment(vertical='top', wrap_text=True)
        cell.border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
        if fill is not None:
            cell.fill = fill



def _autosize(ws) -> None:
    for col_cells in ws.columns:
        letter = get_column_letter(col_cells[0].column)
        max_len = 0
        for cell in col_cells:
            value = '' if cell.value is None else str(cell.value)
            max_len = max(max_len, min(len(value), 50))
        ws.column_dimensions[letter].width = max(12, min(max_len + 4, 40))



def build_excel_report(
    *,
    messages: list[MailItem],
    results: list[CandidateResult],
    remaining_unread: int,
    skipped_by_prefilter: int,
    outbox_dir: Path,
) -> Path:
    wb = Workbook()
    summary_ws = wb.active
    summary_ws.title = '报告总览'
    summary_ws.merge_cells('A1:H1')
    summary_ws['A1'] = '招聘筛查日报'
    _style_title(summary_ws['A1'])
    summary_ws.row_dimensions[1].height = 26

    summary_rows = [
        ('生成时间', datetime.now().strftime('%Y-%m-%d %H:%M:%S')),
        ('本次读取邮件', len(messages)),
        ('筛查通过人数', len(results)),
        ('预筛跳过 LLM', skipped_by_prefilter),
        ('剩余未读', remaining_unread),
    ]
    for idx, (label, value) in enumerate(summary_rows, start=3):
        summary_ws[f'A{idx}'] = label
        summary_ws[f'B{idx}'] = value
        summary_ws[f'A{idx}'].fill = SUBHEADER_FILL
        summary_ws[f'A{idx}'].font = Font(bold=True)
        _style_data_row(summary_ws[idx])

    summary_ws['A10'] = '管理建议'
    summary_ws['A10'].fill = SUBHEADER_FILL
    summary_ws['A10'].font = Font(bold=True)
    summary_ws['B10'] = '优先联系 90 分以上候选人；80-89 分作为候补池；手机号缺失的候选人建议从原始附件补录。'
    _style_data_row(summary_ws[10], fill=WARN_FILL)

    mail_ws = wb.create_sheet('本轮读取名单')
    mail_headers = ['UID', '发件人', '主题']
    mail_ws.append(mail_headers)
    _style_header_row(mail_ws[1])
    for item in messages:
        mail_ws.append([
            item.uid,
            str(item.message.get('from') or ''),
            str(item.message.get('subject') or ''),
        ])
        _style_data_row(mail_ws[mail_ws.max_row], fill=SKIP_FILL)

    pass_ws = wb.create_sheet('通过名单')
    pass_headers = [
        '候选人', '匹配岗位', '评分', '分档', '手机号', '邮箱', '年限(识别)',
        '邮件主题', '发件人', '附件', '摘要', '推荐建议'
    ]
    pass_ws.append(pass_headers)
    _style_header_row(pass_ws[1])
    for result in sorted(results, key=lambda r: (-r.score, r.matched_jd_title, r.candidate_name)):
        mail_meta = result.work_dir / 'mail.json'
        mail_data = {}
        if mail_meta.exists():
            import json
            mail_data = json.loads(mail_meta.read_text(encoding='utf-8'))
        material = (result.work_dir / 'candidate_material.txt').read_text(encoding='utf-8', errors='ignore') if (result.work_dir / 'candidate_material.txt').exists() else ''
        pass_ws.append([
            result.candidate_name,
            result.matched_jd_title,
            result.score,
            result.band,
            _find_mobile(material),
            _find_email(material),
            _find_years(material),
            result.subject,
            result.sender,
            _attachment_names(mail_data.get('documents', [])) if isinstance(mail_data, dict) else '',
            result.summary,
            result.recommendation,
        ])
        _style_data_row(pass_ws[pass_ws.max_row], fill=PASS_FILL)

    summary_by_jd_ws = wb.create_sheet('岗位汇总')
    summary_by_jd_ws.append(['岗位', '通过人数', '90-99', '80-89', '建议'])
    _style_header_row(summary_by_jd_ws[1])
    grouped: dict[str, dict[str, int]] = {}
    for result in results:
        bucket = grouped.setdefault(result.matched_jd_title, {'90-99': 0, '80-89': 0, 'total': 0})
        bucket['total'] += 1
        bucket[result.band] = bucket.get(result.band, 0) + 1
    for jd_title, info in sorted(grouped.items()):
        suggestion = '优先面试' if info.get('90-99', 0) else '建议电话初筛'
        summary_by_jd_ws.append([jd_title, info['total'], info.get('90-99', 0), info.get('80-89', 0), suggestion])
        _style_data_row(summary_by_jd_ws[summary_by_jd_ws.max_row])

    for ws in wb.worksheets:
        ws.freeze_panes = 'A2'
        _autosize(ws)

    file_name = f"interviewer-report-{datetime.now().strftime('%Y-%m-%d-%H%M%S')}.xlsx"
    output_path = outbox_dir / file_name
    wb.save(output_path)
    return output_path
